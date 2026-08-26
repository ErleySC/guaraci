# -*- coding: utf-8 -*-
"""plano_coleta.py -- Distribuicao de amostras entre sessoes de coleta e
ordem de leitura aleatorizada (Bloco 10, `guaraci plan`).

POR QUE ESTE MODULO EXISTE
---------------------------
Duas formas de confundimento ja documentadas neste projeto (achados de
auditorias anteriores, fora do repo -- `~/.guaraci_local/auditoria_
privada/medir_confundimento_data.py`, `medir_ordem_leitura.py`) sao
EVITAVEIS no momento do PLANEJAMENTO da coleta, antes de virarem dado a
corrigir depois:

1. CLASSE x SESSAO: se uma sessao de coleta contem so' uma classe (ou
   uma classe fortemente dominante), qualquer deriva instrumental/
   temporal daquela sessao fica indistinguivel de efeito de classe --
   exatamente o confundimento que motivou `dados_io.session_from_mae_id`
   e a separacao treino/teste por sessao no Bloco 9b, so' que agora
   evitado na COLETA em vez de mitigado depois na analise.
2. ORDEM DE LEITURA x TEOR: se as amostras sao lidas em ordem crescente
   de teor (ou qualquer ordem correlacionada com uma variavel de
   interesse), qualquer deriva do instrumento ao longo da sessao (fonte
   esquentando, purga de CO2/H2O, sujeira acumulando na janela) entra no
   modelo como se fosse sinal quimico -- medido em sessao anterior como
   rho de Spearman proximo de 1 quando a ordem de leitura segue o teor
   dentro do bloco.

Este modulo distribui as amostras por sessao de forma BALANCEADA
(round-robin determinstico -- o objetivo e' balanceamento, nao
aleatoriedade nessa etapa) e ALEATORIZA a ordem de leitura DENTRO de
cada sessao (aqui sim, aleatoriedade e' o ponto).

Integra com `plano_amostral.py` (P1, Bloco 10): aquele modulo decide
QUANTAS amostras sao necessarias (`n_minimum_conformal`/
`ddsimca_sample_size_guidance`); este decide COMO distribui-las
entre sessoes e em que ordem le-las.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np

from guaraci.plano_amostral import (
    n_minimum_conformal,
    ddsimca_sample_size_guidance,
)

__all__ = [
    "CollectionItem",
    "CollectionPlan",
    "plan_collection",
    "plan_from_statistical_target",
    "export_markdown",
    "export_excel",
]


@dataclass
class CollectionItem:
    """Uma AMOSTRA FISICA no plano -- ver alerta sobre replica tecnica em
    `plan_collection`. `ordem_na_sessao` ja' e' a ordem de leitura
    ALEATORIZADA (0-based), nao a ordem de alocacao."""

    classe: str
    replica_idx: int
    sessao: int
    ordem_na_sessao: int


@dataclass
class CollectionPlan:
    classes: List[str]
    n_por_classe: int
    n_sessoes: int
    itens: List[CollectionItem]
    alertas: List[str] = field(default_factory=list)

    def itens_da_sessao(self, sessao: int) -> List[CollectionItem]:
        return sorted((i for i in self.itens if i.sessao == sessao),
                       key=lambda i: i.ordem_na_sessao)


def plan_collection(classes: Sequence[str], n_por_classe: int,
                     n_sessoes: int, *, n_brancos_por_sessao: int = 1,
                     seed: Optional[int] = None) -> CollectionPlan:
    """Distribui `n_por_classe` amostras de cada classe entre `n_sessoes`
    sessoes, balanceado (round-robin), e aleatoriza a ordem de leitura
    DENTRO de cada sessao. Ver docstring do modulo para o motivo dos dois
    confundimentos que isso evita.

    `seed`: None usa entropia do sistema (plano de coleta real -- cada
    geracao deve ser genuinamente aleatoria); um inteiro fixo e' so' para
    reprodutibilidade em teste.
    """
    if n_por_classe < 1:
        raise ValueError("n_por_classe deve ser >= 1")
    if n_sessoes < 1:
        raise ValueError("n_sessoes deve ser >= 1")
    if not classes:
        raise ValueError("classes nao pode ser vazio")

    rng = np.random.default_rng(seed)
    itens: List[CollectionItem] = []

    for classe in classes:
        for replica_idx in range(n_por_classe):
            sessao = replica_idx % n_sessoes
            itens.append(CollectionItem(classe=classe, replica_idx=replica_idx,
                                         sessao=sessao, ordem_na_sessao=-1))

    for sessao in range(n_sessoes):
        idx_sessao = [i for i, item in enumerate(itens)
                       if item.sessao == sessao]
        ordem = rng.permutation(len(idx_sessao))
        for pos, idx in zip(ordem, idx_sessao):
            itens[idx].ordem_na_sessao = int(pos)

    alertas: List[str] = []
    if n_por_classe % n_sessoes != 0:
        resto = n_por_classe % n_sessoes
        alertas.append(
            f"n_por_classe ({n_por_classe}) nao e' multiplo de n_sessoes "
            f"({n_sessoes}): {resto} sessao(oes) recebem 1 amostra a mais "
            "de cada classe (balanceamento residual pequeno -- confira a "
            "tabela por sessao antes de coletar).")
    if n_sessoes == 1:
        alertas.append(
            "APENAS 1 SESSAO: nao ha' como separar efeito de classe de "
            "deriva instrumental/temporal -- o confundimento classe x "
            "sessao fica estruturalmente impossivel de descartar com "
            "este plano. Distribua a coleta em >=2 sessoes se possivel.")
    alertas.append(
        "REPLICA TECNICA != AMOSTRA FISICA: cada item deste plano e' UMA "
        "amostra fisica independente (novo preparo), nao uma releitura do "
        "mesmo vial. Replicas tecnicas (T1/T2/T3 da MESMA amostra) nao "
        "contam como amostras adicionais para fins de tamanho amostral "
        "(ver plano_amostral.py) -- leia cada item quantas vezes o "
        "protocolo do laboratorio pedir, mas isso nao aumenta o `n` "
        "usado para a garantia estatistica.")
    alertas.append(
        "BRANCOS/CONTROLES: inclua pelo menos "
        f"{max(1, n_brancos_por_sessao)} leitura(s) de branco/referencia "
        "por sessao, intercalada(s) ao longo da ordem de leitura (nao so' "
        "no inicio/fim) -- e' o que permite detectar deriva instrumental "
        "depois, sem precisar reler amostras.")

    return CollectionPlan(classes=list(classes), n_por_classe=n_por_classe,
                           n_sessoes=n_sessoes, itens=itens, alertas=alertas)


def plan_from_statistical_target(
        classes: Sequence[str], n_sessoes: int, *,
        alpha_conformal: Optional[float] = None,
        cobertura_ddsimca: Optional[float] = None,
        n_brancos_por_sessao: int = 1,
        seed: Optional[int] = None) -> Tuple[CollectionPlan, Dict[str, Any]]:
    """Combina `plano_amostral.py` (quanto coletar) com `plan_collection`
    (como distribuir/ordenar). Passe exatamente UM alvo estatistico:
    `alpha_conformal` (gate conformal, Identificar/agrupado) OU
    `cobertura_ddsimca` (DD-SIMCA por especie, com o teto de plato ja
    embutido em `ddsimca_sample_size_guidance` -- ver esse modulo
    para o motivo de nao prometer cobertura inalcancavel).

    Retorna (plano, metadados) -- metadados inclui `n_por_classe`,
    `origem` (qual alvo foi usado) e, se DD-SIMCA, a `DDSimcaGuidance`
    completa (para quem quiser inspecionar a ressalva original).
    """
    if (alpha_conformal is None) == (cobertura_ddsimca is None):
        raise ValueError(
            "passe exatamente um de alpha_conformal ou cobertura_ddsimca, "
            "nunca os dois nem nenhum")

    metadados: Dict[str, Any] = {}
    if alpha_conformal is not None:
        n = n_minimum_conformal(alpha_conformal)
        metadados["origem"] = f"conformal (alpha={alpha_conformal})"
        metadados["n_por_classe"] = n
    else:
        orientacao = ddsimca_sample_size_guidance(cobertura_ddsimca)
        if not orientacao.alcancavel:
            raise ValueError(
                f"cobertura-alvo {cobertura_ddsimca} NAO alcancavel via "
                f"DD-SIMCA -- {orientacao.ressalva}")
        n = orientacao.n_sugerido
        metadados["origem"] = f"DD-SIMCA (cobertura={cobertura_ddsimca})"
        metadados["n_por_classe"] = n
        metadados["orientacao_ddsimca"] = orientacao

    plano = plan_collection(classes, n, n_sessoes,
                             n_brancos_por_sessao=n_brancos_por_sessao,
                             seed=seed)
    plano.alertas.insert(
        0, f"n por classe ({n}) calculado via {metadados['origem']}.")
    if "orientacao_ddsimca" in metadados:
        plano.alertas.insert(1, metadados["orientacao_ddsimca"].ressalva)
    return plano, metadados


def export_markdown(plano: CollectionPlan) -> str:
    """Formato PRIMARIO de saida (P4, Bloco 10) -- simples de gerar e
    versionar, sem dependencia nova."""
    linhas = [
        "# Plano de Coleta -- GUARACI", "",
        f"**Classes:** {', '.join(plano.classes)}  ",
        f"**Amostras por classe:** {plano.n_por_classe}  ",
        f"**Sessoes:** {plano.n_sessoes}  ",
        f"**Total de amostras fisicas:** {len(plano.itens)}",
        "", "## Alertas", "",
    ]
    for a in plano.alertas:
        linhas.append(f"- {a}")
    linhas.append("")
    for sessao in range(plano.n_sessoes):
        itens_sessao = plano.itens_da_sessao(sessao)
        linhas.append(f"## Sessao {sessao + 1} ({len(itens_sessao)} amostras)")
        linhas.append("")
        linhas.append("| Ordem de leitura | Classe | Replica |")
        linhas.append("|---|---|---|")
        for item in itens_sessao:
            linhas.append(f"| {item.ordem_na_sessao + 1} | {item.classe} | "
                           f"{item.replica_idx + 1} |")
        linhas.append("")
    return "\n".join(linhas)


def export_excel(plano: CollectionPlan, caminho: str) -> str:
    """Planilha (P4, Bloco 10) -- reaproveita `openpyxl`, ja' dependencia
    do projeto (`reports.py`), nenhuma dependencia nova. 2 abas: ordem de
    leitura completa (para levar ao laboratorio) e alertas."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordem de Leitura"
    ws.append(["Sessao", "Ordem de leitura", "Classe", "Replica"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for item in sorted(plano.itens, key=lambda i: (i.sessao, i.ordem_na_sessao)):
        ws.append([item.sessao + 1, item.ordem_na_sessao + 1, item.classe,
                   item.replica_idx + 1])

    ws_alertas = wb.create_sheet("Alertas")
    ws_alertas.append(["Alerta"])
    ws_alertas["A1"].font = Font(bold=True)
    for a in plano.alertas:
        ws_alertas.append([a])
    ws_alertas.column_dimensions["A"].width = 100
    ws.column_dimensions["C"].width = 20

    wb.save(caminho)
    return caminho
