# -*- coding: utf-8 -*-
"""Testes de guaraci.plano_coleta (Bloco 10).

Os dois invariantes centrais que este arquivo prova (ver docstring do
modulo para o motivo): (1) nenhuma sessao fica dominada por uma unica
classe (confundimento classe x sessao) -- todo balanceamento e'
verificado exatamente, nao so' "parece razoavel"; (2) a ordem de leitura
dentro de cada sessao NAO segue a ordem de alocacao/replica -- e' uma
permutacao de verdade, testada estatisticamente contra o caso trivial
"ordem = ordem de alocacao".
"""
from __future__ import annotations

import os

import openpyxl
import pytest

from guaraci.plano_coleta import (
    CollectionItem,
    export_excel,
    export_markdown,
    plan_from_statistical_target,
    plan_collection,
)


# ── Validacao de entrada ───────────────────────────────────────────────

def test_n_por_classe_invalido_falha_alto():
    with pytest.raises(ValueError, match="n_por_classe"):
        plan_collection(["A", "B"], 0, 2)


def test_n_sessoes_invalido_falha_alto():
    with pytest.raises(ValueError, match="n_sessoes"):
        plan_collection(["A", "B"], 10, 0)


def test_classes_vazio_falha_alto():
    with pytest.raises(ValueError, match="classes"):
        plan_collection([], 10, 2)


# ── Confundimento classe x sessao ────────────────────────────────────────

def test_toda_sessao_recebe_todas_as_classes():
    """O invariante central: nenhuma sessao pode ficar sem uma classe
    (isso seria o confundimento classe x sessao que o modulo existe para
    evitar)."""
    classes = ["Andiroba", "Bacaba", "Coco", "Buriti"]
    plano = plan_collection(classes, n_por_classe=12, n_sessoes=4, seed=0)
    for sessao in range(plano.n_sessoes):
        classes_na_sessao = {i.classe for i in plano.itens_da_sessao(sessao)}
        assert classes_na_sessao == set(classes), (
            f"sessao {sessao} nao tem todas as classes -- confundimento "
            "classe x sessao")


def test_distribuicao_e_balanceada_quando_divisivel():
    classes = ["A", "B", "C"]
    plano = plan_collection(classes, n_por_classe=12, n_sessoes=4, seed=0)
    for sessao in range(plano.n_sessoes):
        for classe in classes:
            n_classe_sessao = sum(
                1 for i in plano.itens_da_sessao(sessao) if i.classe == classe)
            assert n_classe_sessao == 3   # 12/4 exato


def test_total_de_amostras_bate():
    plano = plan_collection(["A", "B"], n_por_classe=7, n_sessoes=3, seed=0)
    assert len(plano.itens) == 7 * 2


def test_desbalanceamento_residual_gera_alerta():
    plano = plan_collection(["A", "B"], n_por_classe=10, n_sessoes=3, seed=0)
    assert any("nao e' multiplo" in a for a in plano.alertas)


def test_uma_sessao_so_gera_alerta_forte():
    plano = plan_collection(["A", "B"], n_por_classe=10, n_sessoes=1, seed=0)
    assert any("APENAS 1 SESSAO" in a for a in plano.alertas)


# ── Ordem de leitura aleatorizada ────────────────────────────────────────

def test_ordem_de_leitura_nao_e_a_ordem_de_alocacao():
    """Contra-prova central: a ordem de leitura tem que ser uma
    PERMUTACAO de verdade, nao a ordem trivial em que os itens foram
    alocados (que seria: todas as replicas de uma classe seguidas, depois
    a proxima classe -- exatamente o padrao que confunde ordem com
    classe/teor)."""
    classes = [f"Especie_{i}" for i in range(10)]
    plano = plan_collection(classes, n_por_classe=5, n_sessoes=1, seed=1)
    itens_sessao = plano.itens_da_sessao(0)
    sequencia_classes = [i.classe for i in itens_sessao]
    # se a ordem fosse a de alocacao, classes apareceriam em blocos
    # (AAAAA BBBBB CCCCC ...) -- conta quantas vezes a classe muda de um
    # item pro proximo; ordem de alocacao teria poucas mudancas (9), uma
    # permutacao de verdade tem muito mais.
    mudancas = sum(1 for a, b in zip(sequencia_classes, sequencia_classes[1:])
                    if a != b)
    assert mudancas > 20, (
        "ordem de leitura parece seguir a ordem de alocacao (poucas "
        "mudancas de classe consecutivas) -- nao esta' embaralhada")


def test_ordem_de_leitura_e_permutacao_completa_por_sessao():
    plano = plan_collection(["A", "B", "C"], n_por_classe=4, n_sessoes=2, seed=2)
    for sessao in range(plano.n_sessoes):
        ordens = sorted(i.ordem_na_sessao for i in plano.itens_da_sessao(sessao))
        assert ordens == list(range(len(ordens))), (
            "ordem_na_sessao tem que ser uma permutacao 0..n-1 sem "
            "buracos nem repeticao")


def test_seed_fixa_e_reprodutivel_seed_none_nao_repete():
    p1 = plan_collection(["A", "B"], 6, 2, seed=42)
    p2 = plan_collection(["A", "B"], 6, 2, seed=42)
    ordem1 = [(i.classe, i.replica_idx, i.sessao, i.ordem_na_sessao)
              for i in p1.itens]
    ordem2 = [(i.classe, i.replica_idx, i.sessao, i.ordem_na_sessao)
              for i in p2.itens]
    assert ordem1 == ordem2


# ── Alertas sempre presentes ──────────────────────────────────────────

def test_alerta_replica_tecnica_sempre_presente():
    plano = plan_collection(["A"], 5, 2, seed=0)
    assert any("REPLICA TECNICA" in a for a in plano.alertas)


def test_alerta_brancos_sempre_presente():
    plano = plan_collection(["A"], 5, 2, seed=0)
    assert any("BRANCOS" in a for a in plano.alertas)


# ── Integracao com plano_amostral.py ─────────────────────────────────

def test_planejar_a_partir_de_alfa_conformal_usa_n_minimum_for_alpha():
    from guaraci.conformal import n_minimum_for_alpha
    plano, meta = plan_from_statistical_target(
        ["A", "B"], n_sessoes=2, alpha_conformal=0.10, seed=0)
    assert meta["n_por_classe"] == n_minimum_for_alpha(0.10)
    assert plano.n_por_classe == n_minimum_for_alpha(0.10)


def test_planejar_a_partir_de_cobertura_ddsimca_alcancavel():
    plano, meta = plan_from_statistical_target(
        ["A", "B"], n_sessoes=2, cobertura_ddsimca=0.90, seed=0)
    assert meta["n_por_classe"] == 20   # ver tabela medida em plano_amostral
    assert "orientacao_ddsimca" in meta


def test_planejar_a_partir_de_cobertura_ddsimca_inalcancavel_falha_alto():
    """Contra-prova D4-like: nunca gera um plano prometendo uma
    cobertura-alvo que o DD-SIMCA nao sustenta (mesmo espirito do gate
    de Quantificar no Bloco 9b -- nunca fingir garantia que nao existe)."""
    with pytest.raises(ValueError, match="NAO alcancavel"):
        plan_from_statistical_target(
            ["A", "B"], n_sessoes=2, cobertura_ddsimca=0.99, seed=0)


def test_planejar_a_partir_de_alvo_exige_exatamente_um():
    with pytest.raises(ValueError, match="exatamente um"):
        plan_from_statistical_target(["A"], n_sessoes=1)
    with pytest.raises(ValueError, match="exatamente um"):
        plan_from_statistical_target(
            ["A"], n_sessoes=1, alpha_conformal=0.05, cobertura_ddsimca=0.9)


# ── Exportacao ─────────────────────────────────────────────────────────

def test_exportar_markdown_contem_todas_as_sessoes_e_alertas():
    plano = plan_collection(["A", "B"], 6, 2, seed=0)
    md = export_markdown(plano)
    assert "Sessao 1" in md and "Sessao 2" in md
    for a in plano.alertas:
        assert a in md


def test_exportar_excel_gera_arquivo_com_2_abas(tmp_path):
    plano = plan_collection(["A", "B", "C"], 4, 2, seed=0)
    caminho = str(tmp_path / "plano.xlsx")
    export_excel(plano, caminho)
    assert os.path.isfile(caminho)

    wb = openpyxl.load_workbook(caminho)
    assert set(wb.sheetnames) == {"Ordem de Leitura", "Alertas"}
    ws = wb["Ordem de Leitura"]
    # cabecalho + 1 linha por item
    assert ws.max_row == 1 + len(plano.itens)
    ws_alertas = wb["Alertas"]
    assert ws_alertas.max_row == 1 + len(plano.alertas)


def test_itens_da_sessao_e_um_item_de_dataclass_completo():
    plano = plan_collection(["A"], 3, 1, seed=0)
    item = plano.itens_da_sessao(0)[0]
    assert isinstance(item, CollectionItem)
    assert item.classe == "A"
    assert 0 <= item.replica_idx < 3


# =========================================================================
#  D6 (mesmo padrao do menu B, Bloco 9b): CLI (menu J, Planejamento de
#  Coleta) fim a fim, sem mock de logica cientifica.
# =========================================================================

def test_menu_plan_cli_end_to_end_conformal(monkeypatch, tmp_path):
    import guaraci.guaraci as guaraci_mod

    prefixo = str(tmp_path / "plano_teste")
    respostas = iter([
        "Andiroba, Bacaba, Coco",   # classes
        "2",                          # n_sessoes
        "C",                          # alvo: conformal
        "0.10",                       # alpha
        prefixo,                      # prefixo de saida
        "",                           # Enter no _pause() final
    ])
    monkeypatch.setattr("builtins.input", lambda *a, **k: next(respostas))

    guaraci_mod._menu_plan(guaraci_mod.Config())

    assert os.path.isfile(prefixo + ".md")
    assert os.path.isfile(prefixo + ".xlsx")
    texto_md = open(prefixo + ".md", encoding="utf-8").read()
    assert "Andiroba" in texto_md and "Bacaba" in texto_md and "Coco" in texto_md
    assert "Sessao 1" in texto_md and "Sessao 2" in texto_md


def test_menu_plan_cli_cobertura_ddsimca_inalcancavel_nao_gera_arquivo(
        monkeypatch, tmp_path):
    """Contra-prova via CLI: pedir uma cobertura-alvo acima do plato
    medido nunca produz um plano/arquivo -- a mensagem de erro aparece e
    a funcao retorna sem escrever nada (nunca promete o inalcancavel)."""
    import guaraci.guaraci as guaraci_mod

    prefixo = str(tmp_path / "plano_impossivel")
    respostas = iter([
        "Andiroba, Bacaba",
        "2",
        "D",       # alvo: DD-SIMCA
        "0.99",    # acima do plato (~0.94) -- inalcancavel
        "",        # _pause() apos o erro
    ])
    monkeypatch.setattr("builtins.input", lambda *a, **k: next(respostas))

    guaraci_mod._menu_plan(guaraci_mod.Config())

    assert not os.path.isfile(prefixo + ".md")
    assert not os.path.isfile(prefixo + ".xlsx")
