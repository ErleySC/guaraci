# -*- coding: utf-8 -*-
"""Contrato de COLUNAS DE SAIDA (Passo 89) -- dívida registrada desde o
Bloco B: o teste de contrato de API pública (`test_contrato_api_publica.py`)
cobre assinaturas/dataclasses, mas explicitamente NÃO cobre nomes de coluna
de CSV/Excel gerados dinamicamente (ver docstring daquele arquivo) — essa é
a superfície mais provável de ser consumida por um script de usuário
automatizando algo em cima do GUARACI (`pd.read_csv(...)["Bal.Acc mean"]`
quebra em silêncio se a coluna virar "BalAcc_mean").

MECANISMO: mesmo padrão de `test_contrato_api_publica.py` -- snapshot
golden em `tests/golden/contrato_saida_tabular.json`, regravado com
`GUARACI_REGRAVAR_GOLDEN=1 pytest tests/test_contrato_saida_tabular.py`.
Cada entrada é capturada por EXECUÇÃO REAL contra dado sintético (nunca
lista digitada à mão) -- não existe objeto Python introspectável que
declare essas colunas sem rodar a função e inspecionar o resultado.

ESCOPO -- levantado por leitura direta de todo `.to_csv(`/`.to_excel(`/
`Workbook(` em `src/guaraci/` (comando: `grep -rn "to_csv|to_excel|Workbook(" src/guaraci/`):
- `reports.py::generate_excel_report` NÃO tem entrada própria aqui:
  verificado por leitura (`_preencher_df`/`_cabecalho`) que ele copia
  `list(df.columns)` VERBATIM dos CSVs já cobertos abaixo -- não pode
  renomear uma coluna independentemente da fonte, então protegê-los já
  protege a reempacotação em Excel.
- `auditoria_delineamento.py`/`sentinela_deriva.py` verificados (mesmo
  grep): não produzem CSV/Excel próprio, só dict/relatório consumido em
  memória -- fora do escopo por não existir superfície tabular.
- `linearity.py`/`robustness.py` (Bloco 13d): mesma situação -- o
  resultado vira addendum em `model_card.md` (Markdown, não tabular),
  via `append_linearity_robustness_model_card`.
"""
from __future__ import annotations

import json
import os
import pathlib

import numpy as np
import pandas as pd
import pytest

_GOLDEN = pathlib.Path(__file__).parent / "golden" / "contrato_saida_tabular.json"


# =========================================================================
#  Geradores de dado sintetico (pequenos, so' para exercitar as funcoes --
#  nao sao testes de qualidade de resultado, so' de FORMA da saida).
# =========================================================================

def _dados_classificacao(seed=0, n_grupos_por_classe=6, n_replicas=3, p=25, n_classes=3):
    from sklearn.preprocessing import LabelBinarizer
    rng = np.random.default_rng(seed)
    X_list, y_list, grp_list = [], [], []
    for c in range(n_classes):
        centro_classe = rng.normal(loc=c * 5.0, size=p)
        for g in range(n_grupos_por_classe):
            centro_grupo = centro_classe + rng.normal(scale=0.3, size=p)
            grupo_id = f"C{c}G{g}"
            for _r in range(n_replicas):
                X_list.append(centro_grupo + rng.normal(scale=0.05, size=p))
                y_list.append(c)
                grp_list.append(grupo_id)
    X = np.array(X_list)
    y_int = np.array(y_list)
    grupos = np.array(grp_list)
    lb = LabelBinarizer().fit(y_int)
    return X, y_int, grupos, lb


def _dados_regressao(seed=0, n_especies=2, n_pontos=8, n_replicas=3, p=30):
    rng = np.random.default_rng(seed)
    X_list, conc_list, rot_list, mae_list = [], [], [], []
    especies = [f"Esp_{chr(65 + i)}" for i in range(n_especies)]
    for e_idx, especie in enumerate(especies):
        w_true = rng.normal(size=p)
        centro = rng.normal(loc=e_idx * 6.0, size=p)
        concs = np.linspace(0, 40, n_pontos)
        for p_idx, c in enumerate(concs):
            espectro_base = centro + w_true * (c / 40.0) * 3.0
            grupo_id = f"{especie}_{p_idx:02d}"
            for _r in range(n_replicas):
                X_list.append(espectro_base + rng.normal(scale=0.10, size=p))
                conc_list.append(c)
                rot_list.append(especie)
                mae_list.append(grupo_id)
    X = np.array(X_list)
    conc = np.array(conc_list)
    rotulos = np.array(rot_list)
    mae_id = np.array(mae_list)
    classes_unicas = np.array(especies)
    return X, conc, rotulos, mae_id, classes_unicas


# =========================================================================
#  Captura de colunas -- uma funcao por SAIDA TABULAR, sempre por execucao
#  real (nunca lista escrita a mao).
# =========================================================================

def _colunas_identifiers(pq, tmp_path):
    n = 12
    rng = np.random.default_rng(0)
    rotulos = np.array(["A", "B"] * (n // 2))
    pred_lab = np.array(["A", "B"] * (n // 2))
    scores_pls = rng.normal(size=(n, 3))
    T2 = rng.uniform(0, 5, size=n)
    Q = rng.uniform(0, 5, size=n)
    pasta = str(tmp_path / "identifiers")
    os.makedirs(pasta, exist_ok=True)
    pq.save_identifiers(rotulos, pred_lab, scores_pls, T2, Q, 3.0, 3.0, pasta)
    df = pd.read_csv(os.path.join(pasta, "amostras_identificadores.csv"), sep=";")
    return list(df.columns)


def _colunas_metadados_sanitizados(pq):
    df_meta = pd.DataFrame({
        "titulo": ["amostra1.dx", "amostra2.dx"],
        "arquivo": ["c:/dados/amostra1.dx", "c:/dados/amostra2.dx"],
        "mae_id": ["G1", "G2"],
        "data": ["2026-01-01", "2026-01-02"],
    })
    return list(pq.sanitizar_metadados(df_meta).columns)


def _colunas_benchmark_classifiers(pq, tmp_path):
    X, y_int, grupos, lb = _dados_classificacao()
    cfg = pq.Config(n_splits_cv=3, seed=0, run_shap=False)
    pasta = str(tmp_path / "bench_clf")
    os.makedirs(os.path.join(pasta, pq.NOME_TABELAS), exist_ok=True)
    os.makedirs(os.path.join(pasta, pq.NOME_GRAFICOS), exist_ok=True)
    df = pq.benchmark_classifiers(X, y_int, grupos, lb, n_opt=2, cfg=cfg, pasta=pasta)
    return list(df.columns)


def _colunas_monte_carlo_cv(pq, tmp_path):
    X, y_int, grupos, lb = _dados_classificacao()
    cfg = pq.Config(n_splits_cv=3, seed=0, run_shap=False,
                    n_monte_carlo=3, monte_carlo_include_all=False)
    pasta = str(tmp_path / "monte_carlo")
    os.makedirs(os.path.join(pasta, pq.NOME_TABELAS), exist_ok=True)
    os.makedirs(os.path.join(pasta, pq.NOME_GRAFICOS), exist_ok=True)
    df = pq.monte_carlo_cv(X, y_int, grupos, lb, n_opt=2, cfg=cfg, pasta=pasta)
    return list(df.columns)


def _colunas_benchmark_regression_by_species(pq, tmp_path):
    X, conc, rotulos, mae_id, classes_unicas = _dados_regressao()
    cfg = pq.Config(seed=0, max_lvs=5, frac_cal=0.7)
    pasta = str(tmp_path / "bench_reg")
    os.makedirs(os.path.join(pasta, pq.NOME_TABELAS), exist_ok=True)
    os.makedirs(os.path.join(pasta, pq.NOME_GRAFICOS), exist_ok=True)
    reg_esp = pq.pls_regression_by_species(
        X, conc, rotulos, mae_id, classes_unicas, cfg, pasta, n_splits=3)
    df = pq.benchmark_regression_by_species(
        X, conc, rotulos, mae_id, classes_unicas, cfg, pasta, reg_esp)
    return list(df.columns)


def _colunas_etapa4(pq, tmp_path):
    """Etapa 4 escreve VARIOS CSVs numa so' chamada (ipls/spa/ag/tabela
    final) -- captura os 4."""
    from sklearn.model_selection import StratifiedKFold

    rng = np.random.default_rng(4)
    n, p = 80, 30
    classes = rng.integers(0, 2, size=n)
    X = rng.normal(size=(n, p))
    vars_informativas = (5, 6, 7, 20, 21)
    for i in range(n):
        if classes[i] == 1:
            X[i, list(vars_informativas)] += 3.0
    Y_bin = np.zeros((n, 2))
    for i, c in enumerate(classes):
        Y_bin[i, c] = 1
    cv = StratifiedKFold(n_splits=3, shuffle=True, random_state=4)
    cv_indices = list(cv.split(X, classes))
    wavenumbers = np.linspace(4000, 400, p)

    cfg = pq.Config(run_spa=True, executar_ag=True,
                    spa_n_vars_max=6, spa_n_starts=6,
                    ag_tam_populacao=8, ag_n_geracoes=3, seed=4)
    pasta_dados = str(tmp_path / "etapa4")
    os.makedirs(pasta_dados, exist_ok=True)
    pq.etapa4_selecao_variaveis(
        X, Y_bin, classes, wavenumbers, cv_indices, n_lv=2,
        cfg=cfg, pasta=pasta_dados, pasta_dados=pasta_dados)

    saida = {}
    for nome_arquivo, chave in [
        ("etapa4_ipls_intervalos.csv", "ipls_intervalos"),
        ("etapa4_spa_cadeias.csv", "spa_cadeias"),
        ("etapa4_ag_historico.csv", "ag_historico"),
        ("etapa4_selecao_variaveis.csv", "tabela_final"),
    ]:
        caminho = os.path.join(pasta_dados, nome_arquivo)
        if os.path.isfile(caminho):
            df = pd.read_csv(caminho, sep=";")
            saida[chave] = list(df.columns)
        else:
            saida[chave] = None   # arquivo condicional (ex.: SPA sem cadeia valida)
    return saida


def _colunas_plano_coleta_excel(pq, tmp_path):
    import guaraci.plano_coleta as plano_mod
    plano, _meta = plano_mod.plan_from_statistical_target(
        ["Especie1", "Especie2"], n_sessoes=2, alpha_conformal=0.10)
    caminho = str(tmp_path / "plano.xlsx")
    plano_mod.export_excel(plano, caminho)

    import openpyxl
    wb = openpyxl.load_workbook(caminho)
    return {aba: [c.value for c in wb[aba][1]] for aba in wb.sheetnames}


def _colunas_selecao_amostras_cli(pq, tmp_path, monkeypatch):
    import guaraci.guaraci as guaraci_mod

    caminho_csv = str(tmp_path / "espectros.csv")
    rng = np.random.default_rng(0)
    df = pd.DataFrame({f"canal_{i}": rng.normal(size=20) for i in range(6)})
    df.to_csv(caminho_csv, index=False)
    caminho_saida = str(tmp_path / "saida_selecao.csv")

    respostas = iter([caminho_csv, "", "1", "0.6", caminho_saida, ""])
    monkeypatch.setattr("builtins.input", lambda *a, **k: next(respostas))
    guaraci_mod._menu_selecao_amostras(guaraci_mod.Config())

    return list(pd.read_csv(caminho_saida).columns)


def _colunas_pipeline_run_completo(pq, tmp_path):
    """UMA execucao real de executar() com run_martens=True e
    comparar_pipelines=True -- as duas produzem CSV a partir de um dict
    literal CONSTRUIDO INLINE dentro de executar() (nao atras de uma
    funcao publica dedicada), entao so' rodar de verdade protege contra
    renomear a chave do dict ali. Tambem devolve o caminho do modelo
    salvo (para `predict_samples`, que precisa de um pacote real)."""
    from conftest import achar_pastas_run

    cfg = pq.Config(
        input_folder=str(tmp_path / "dados"), output_root_folder=str(tmp_path / "saida"),
        mode="sintetico", n_per_class=10, n_synthetic_points=60,
        wn_min=400.0, wn_max=4001.0,
        n_splits_cv=3, n_repeats_cv=1, n_permutations=5,
        n_permutations_wold=5, n_bootstrap_vip=3, n_bootstrap_bca=20,
        n_monte_carlo=3, max_lvs=5,
        run_martens=True, comparar_pipelines=True,
    )
    os.makedirs(cfg.input_folder, exist_ok=True)
    pq.executar(cfg)

    runs = achar_pastas_run(tmp_path / "saida")
    assert runs, "executar() nao criou pasta de saida"
    run_dir = pathlib.Path(runs[0])

    saida = {}
    cam_martens = run_dir / pq.NOME_TABELAS / "teste_martens.csv"
    saida["teste_martens"] = (
        list(pd.read_csv(cam_martens, sep=";").columns) if cam_martens.is_file() else None)
    cam_comp = run_dir / pq.NOME_TABELAS / "comparacao_pipelines.csv"
    saida["comparacao_pipelines"] = (
        list(pd.read_csv(cam_comp, sep=";").columns) if cam_comp.is_file() else None)
    cam_ident = run_dir / pq.NOME_TABELAS / "amostras_identificadores.csv"
    saida["identifiers_via_executar"] = (
        list(pd.read_csv(cam_ident, sep=";").columns) if cam_ident.is_file() else None)

    cam_modelo = run_dir / pq.NOME_MODELOS / "modelo_plsda.joblib"
    return saida, (str(cam_modelo) if cam_modelo.is_file() else None)


def _colunas_predict_samples(pq, caminho_modelo):
    if caminho_modelo is None:
        return None
    import guaraci.predicao as pred_mod
    pkg = pred_mod.load_model(caminho_modelo, confiar=True)
    rng = np.random.default_rng(0)
    wn = np.asarray(pkg["wavenumbers"], dtype=float)
    X_new = rng.normal(size=(3, len(wn))) + 1.0
    df = pred_mod.predict_samples(pkg, X_new, wn)
    return list(df.columns)


# =========================================================================
#  Snapshot completo
# =========================================================================

def _gerar_snapshot(pq, tmp_path, monkeypatch) -> dict:
    saida_pipeline, caminho_modelo = _colunas_pipeline_run_completo(pq, tmp_path)
    snapshot = {
        "identifiers_csv": _colunas_identifiers(pq, tmp_path),
        "metadados_sanitizados_csv": _colunas_metadados_sanitizados(pq),
        "benchmark_classifiers_csv": _colunas_benchmark_classifiers(pq, tmp_path),
        "monte_carlo_cv_csv": _colunas_monte_carlo_cv(pq, tmp_path),
        "benchmark_regression_by_species_csv":
            _colunas_benchmark_regression_by_species(pq, tmp_path),
        "etapa4": _colunas_etapa4(pq, tmp_path),
        "plano_coleta_excel": _colunas_plano_coleta_excel(pq, tmp_path),
        "selecao_amostras_csv": _colunas_selecao_amostras_cli(pq, tmp_path, monkeypatch),
        "predict_samples_csv": _colunas_predict_samples(pq, caminho_modelo),
        **saida_pipeline,
    }
    return snapshot


@pytest.fixture(scope="module")
def snapshot_atual(pq, tmp_path_factory, monkeypatch_module):
    tmp_path = tmp_path_factory.mktemp("contrato_saida_tabular")
    return _gerar_snapshot(pq, tmp_path, monkeypatch_module)


@pytest.fixture(scope="module")
def monkeypatch_module():
    from _pytest.monkeypatch import MonkeyPatch
    mp = MonkeyPatch()
    yield mp
    mp.undo()


def _diferencas(esperado: dict, obtido: dict) -> list:
    """Detector de divergencia -- usado pelo teste principal E pela
    contra-prova abaixo (mesma funcao, nao duas implementacoes que
    poderiam divergir uma da outra)."""
    diffs = []
    chaves = set(esperado) | set(obtido)
    for chave in sorted(chaves):
        if esperado.get(chave) != obtido.get(chave):
            diffs.append(
                f"  '{chave}': esperado={esperado.get(chave)!r} "
                f"obtido={obtido.get(chave)!r}")
    return diffs


@pytest.mark.slow
def test_contrato_colunas_de_saida_nao_mudou_sem_intencao(snapshot_atual):
    if os.environ.get("GUARACI_REGRAVAR_GOLDEN") == "1":
        _GOLDEN.parent.mkdir(parents=True, exist_ok=True)
        _GOLDEN.write_text(
            json.dumps(snapshot_atual, indent=2, sort_keys=True,
                       ensure_ascii=False) + "\n",
            encoding="utf-8")
        pytest.skip(f"golden regravado em {_GOLDEN} (GUARACI_REGRAVAR_GOLDEN=1)")

    if not _GOLDEN.exists():
        pytest.fail(
            f"Golden nao existe em {_GOLDEN}. Gere com:\n"
            "  GUARACI_REGRAVAR_GOLDEN=1 pytest tests/test_contrato_saida_tabular.py")

    esperado = json.loads(_GOLDEN.read_text(encoding="utf-8"))
    diffs = _diferencas(esperado, snapshot_atual)
    if diffs:
        pytest.fail(
            "Contrato de colunas de saida mudou:\n" + "\n".join(diffs) +
            "\n\nSe a mudanca foi INTENCIONAL, regrave com:\n"
            "  GUARACI_REGRAVAR_GOLDEN=1 pytest tests/test_contrato_saida_tabular.py")


# =========================================================================
#  Contra-prova: simula uma mudanca de nome de coluna deliberada em uma
#  saida REAL (monkeypatch de `save_identifiers`, produzindo um CSV com
#  "classe_predita" renomeada para "classe_pred") e confirma que
#  `_diferencas` -- o MESMO detector usado pelo teste principal, nao uma
#  segunda implementacao -- acusa a mudanca. Sem isto, o teste principal
#  poderia estar passando por vacuidade (sempre "sem diferenca",
#  indiferente ao que de fato mudou).
# =========================================================================

def test_contraprova_renomear_coluna_de_verdade_faz_o_detector_acusar(
        pq, tmp_path, monkeypatch):
    colunas_originais = _colunas_identifiers(pq, tmp_path / "original")

    import guaraci.resultados_io as resultados_io_mod
    _to_csv_original = pd.DataFrame.to_csv

    def _to_csv_com_coluna_renomeada(self, *args, **kwargs):
        if "classe_predita" in self.columns:
            self = self.rename(columns={"classe_predita": "classe_pred"})
        return _to_csv_original(self, *args, **kwargs)

    monkeypatch.setattr(resultados_io_mod.pd.DataFrame, "to_csv",
                        _to_csv_com_coluna_renomeada)
    colunas_com_bug_simulado = _colunas_identifiers(pq, tmp_path / "renomeado")

    assert colunas_com_bug_simulado != colunas_originais, (
        "o monkeypatch deveria ter mudado a coluna 'classe_predita' -> "
        "'classe_pred' -- se as colunas continuam identicas, a simulacao "
        "nao alterou nada e a contra-prova nao prova o que deveria")

    diffs = _diferencas(
        {"identifiers_csv": colunas_originais},
        {"identifiers_csv": colunas_com_bug_simulado})
    assert diffs, (
        "o detector (_diferencas, o MESMO usado pelo teste principal) "
        "deveria acusar a renomeacao de coluna simulada -- se nao "
        "acusou, o teste principal nao protege nada de verdade contra "
        "esta classe de regressao")
